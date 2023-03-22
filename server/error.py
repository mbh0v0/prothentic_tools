import os
from openpyxl import Workbook
import xlrd

current_directory = os.path.dirname(os.path.abspath(__file__))
log_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "request_log.xlsx")
time_log_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "time_log.xlsx")


def get_error(message):
    if "Errno 11001" in str(message):
        return "主机ip填写错误"
    elif "[Errno None]" in str(message):
        return "端口号填写错误"
    elif "Authentication failed" in str(message):
        return "用户名或密码错误"
    else:
        return str(message)


def get_server(log_name):
    if "admin-management" in log_name:
        return "admin-management"
    elif "clickhouse-api" in log_name:
        return "clickhouse-api"
    elif "etl" in log_name:
        return "etl"
    elif "process-analysis" in log_name:
        return "process-analysis"
    elif "resource-portal" in log_name:
        return "resource-portal"
    elif "social-network" in log_name:
        return "social-network"
    elif "spring-gateway" in log_name:
        return "spring-gateway"


def set_request_log(log_list):
    workbook = Workbook()
    sheet = workbook.active
    for i in range(len(log_list)):
        for j in range(4):
            sheet.cell(i + 1, j + 1, log_list[i][j])
    workbook.save(log_path)


def set_time_log(log_list):
    workbook = Workbook()
    sheet = workbook.active
    for i in range(len(log_list)):
        for j in range(4):
            sheet.cell(i + 1, j + 1, log_list[i][j])
    workbook.save(time_log_path)


def get_request_log():
    workbook = xlrd.open_workbook(log_path)
    worksheet = workbook.sheet_by_index(0)
    all_list = []
    if worksheet.nrows == 0:
        return None
    for i in range(worksheet.nrows):
        time_list = [worksheet.row(i)[0].value, worksheet.row(i)[1].value,
                     worksheet.row(i)[2].value, worksheet.row(i)[3].value]
        all_list.append(time_list)
    return all_list


def get_time_log():
    workbook = xlrd.open_workbook(time_log_path)
    worksheet = workbook.sheet_by_index(0)
    all_list = []
    if worksheet.nrows == 0:
        return None
    for i in range(worksheet.nrows):
        time_list = [worksheet.row(i)[0].value, worksheet.row(i)[1].value,
                     worksheet.row(i)[2].value, worksheet.row(i)[3].value]
        all_list.append(time_list)
    return all_list
